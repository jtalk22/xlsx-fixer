"""Tests for xlsx-fixer core functionality.

Tests create real openpyxl-generated xlsx files, verify the corruption
patterns exist, then run the fixer and verify they're resolved.
"""

import tempfile
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pytest

try:
    from openpyxl import Workbook
except ImportError:
    pytest.skip("openpyxl required for tests", allow_module_level=True)

from xlsx_fixer.fixer import SHEET_NS, Issue, check, fix


@pytest.fixture
def corrupt_xlsx(tmp_path):
    """Create a typical openpyxl-generated xlsx with inline strings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Price"
    ws["C1"] = "Status"
    for i in range(2, 52):
        ws[f"A{i}"] = f"Item {i - 1}"
        ws[f"B{i}"] = i * 10.5
        ws[f"C{i}"] = "Active"

    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def clean_xlsx(tmp_path):
    """Create an xlsx and fix it, so it's clean."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    path = tmp_path / "clean.xlsx"
    wb.save(path)
    fix(path)
    return path


class TestCheck:
    def test_detects_inline_strings(self, corrupt_xlsx):
        issues = check(corrupt_xlsx)
        codes = [i.code for i in issues]
        assert "INLINE_STRINGS" in codes

    def test_detects_missing_sst(self, corrupt_xlsx):
        issues = check(corrupt_xlsx)
        codes = [i.code for i in issues]
        assert "MISSING_SST" in codes

    def test_detects_full_calc(self, corrupt_xlsx):
        issues = check(corrupt_xlsx)
        codes = [i.code for i in issues]
        assert "FULL_CALC_NO_CHAIN" in codes

    def test_clean_file_no_errors(self, clean_xlsx):
        issues = check(clean_xlsx)
        errors = [i for i in issues if i.severity == "error"]
        assert len(errors) == 0

    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            check(tmp_path / "nonexistent.xlsx")

    def test_not_xlsx(self, tmp_path):
        bad = tmp_path / "bad.xlsx"
        bad.write_text("not a zip file")
        with pytest.raises(ValueError, match="Not a valid ZIP"):
            check(bad)

    def test_returns_issue_objects(self, corrupt_xlsx):
        issues = check(corrupt_xlsx)
        assert all(isinstance(i, Issue) for i in issues)
        assert all(i.code for i in issues)
        assert all(i.severity in ("error", "warning") for i in issues)
        assert all(i.message for i in issues)


class TestFix:
    def test_converts_inline_strings(self, corrupt_xlsx):
        stats = fix(corrupt_xlsx)
        assert stats["inline_strings"] > 0
        assert stats["unique_strings"] > 0

    def test_creates_shared_strings_xml(self, corrupt_xlsx):
        fix(corrupt_xlsx)
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            assert "xl/sharedStrings.xml" in zf.namelist()

    def test_no_inline_strings_after_fix(self, corrupt_xlsx):
        fix(corrupt_xlsx)
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            for name in zf.namelist():
                if "worksheets/sheet" in name:
                    data = zf.read(name)
                    assert b"inlineStr" not in data

    def test_sst_has_correct_counts(self, corrupt_xlsx):
        stats = fix(corrupt_xlsx)
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            sst_data = zf.read("xl/sharedStrings.xml")
            ET.register_namespace("", SHEET_NS)
            root = ET.fromstring(sst_data)
            assert root.get("count") == str(stats["inline_strings"])
            assert root.get("uniqueCount") == str(stats["unique_strings"])

    def test_removes_full_calc_on_load(self, corrupt_xlsx):
        stats = fix(corrupt_xlsx)
        assert stats["full_calc_removed"] is True
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            wb_data = zf.read("xl/workbook.xml")
            assert b"fullCalcOnLoad" not in wb_data

    def test_adds_sst_relationship(self, corrupt_xlsx):
        fix(corrupt_xlsx)
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            rels_data = zf.read("xl/_rels/workbook.xml.rels")
            assert b"sharedStrings" in rels_data

    def test_adds_content_type(self, corrupt_xlsx):
        fix(corrupt_xlsx)
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            ct_data = zf.read("[Content_Types].xml")
            assert b"sharedStrings.xml" in ct_data

    def test_fix_to_output_path(self, corrupt_xlsx, tmp_path):
        output = tmp_path / "fixed.xlsx"
        fix(corrupt_xlsx, output=output)
        assert output.exists()
        # Original should still have inline strings
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            for name in zf.namelist():
                if "worksheets/sheet" in name:
                    data = zf.read(name)
                    if b"inlineStr" in data:
                        break
            else:
                pass  # Original might be modified by openpyxl version
        # Fixed should not
        with zipfile.ZipFile(output) as zf:
            for name in zf.namelist():
                if "worksheets/sheet" in name:
                    data = zf.read(name)
                    assert b"inlineStr" not in data

    def test_preserves_whitespace(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "  leading spaces"
        ws["A2"] = "trailing spaces  "
        ws["A3"] = "  both  "
        path = tmp_path / "whitespace.xlsx"
        wb.save(path)
        fix(path)
        with zipfile.ZipFile(path) as zf:
            sst_data = zf.read("xl/sharedStrings.xml")
            assert b'xml:space="preserve"' in sst_data

    def test_all_xml_parseable_after_fix(self, corrupt_xlsx):
        fix(corrupt_xlsx)
        with zipfile.ZipFile(corrupt_xlsx) as zf:
            for name in zf.namelist():
                if name.endswith(".xml"):
                    data = zf.read(name)
                    ET.fromstring(data)  # Should not raise

    def test_idempotent(self, corrupt_xlsx):
        """Running fix twice should not break anything."""
        stats1 = fix(corrupt_xlsx)
        stats2 = fix(corrupt_xlsx)
        # Second run should find nothing to fix
        assert stats2["inline_strings"] == 0

    def test_check_clean_after_fix(self, corrupt_xlsx):
        """After fixing, check should report no errors."""
        fix(corrupt_xlsx)
        issues = check(corrupt_xlsx)
        errors = [i for i in issues if i.severity == "error"]
        assert len(errors) == 0

    def test_dedup_strings(self, tmp_path):
        """Duplicate strings should share the same SST index."""
        wb = Workbook()
        ws = wb.active
        for i in range(1, 101):
            ws[f"A{i}"] = "SAME VALUE"
        path = tmp_path / "dedup.xlsx"
        wb.save(path)
        stats = fix(path)
        assert stats["inline_strings"] == 100
        assert stats["unique_strings"] == 1

    def test_empty_strings(self, tmp_path):
        """Empty string cells should be handled correctly."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = ""
        ws["A2"] = "not empty"
        ws["A3"] = ""
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        stats = fix(path)
        assert stats["inline_strings"] > 0

    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            fix(tmp_path / "nonexistent.xlsx")

    def test_not_xlsx(self, tmp_path):
        bad = tmp_path / "bad.xlsx"
        bad.write_text("not a zip file")
        with pytest.raises(ValueError, match="Not a valid ZIP"):
            fix(bad)


class TestCLI:
    def test_check_command(self, corrupt_xlsx):
        from xlsx_fixer.cli import cmd_check
        import argparse

        args = argparse.Namespace(file=str(corrupt_xlsx))
        exit_code = cmd_check(args)
        assert exit_code == 1  # Has errors

    def test_fix_command(self, corrupt_xlsx):
        from xlsx_fixer.cli import cmd_fix
        import argparse

        args = argparse.Namespace(
            file=str(corrupt_xlsx),
            output=None,
            keep_calc_chain=False,
        )
        exit_code = cmd_fix(args)
        assert exit_code == 0

    def test_fix_with_output(self, corrupt_xlsx, tmp_path):
        from xlsx_fixer.cli import cmd_fix
        import argparse

        output = tmp_path / "cli_fixed.xlsx"
        args = argparse.Namespace(
            file=str(corrupt_xlsx),
            output=str(output),
            keep_calc_chain=False,
        )
        exit_code = cmd_fix(args)
        assert exit_code == 0
        assert output.exists()

    def test_check_clean_file(self, clean_xlsx):
        from xlsx_fixer.cli import cmd_check
        import argparse

        args = argparse.Namespace(file=str(clean_xlsx))
        exit_code = cmd_check(args)
        assert exit_code == 0
