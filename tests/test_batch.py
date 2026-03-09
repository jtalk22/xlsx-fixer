"""Tests for batch processing and CLI batch command."""

import argparse
from pathlib import Path
from unittest.mock import patch

import pytest

try:
    from openpyxl import Workbook
except ImportError:
    pytest.skip("openpyxl required for tests", allow_module_level=True)

from xlsx_fixer.fixer import fix_batch


@pytest.fixture
def xlsx_dir(tmp_path):
    """Create a directory with 3 corrupt xlsx files."""
    for i in range(3):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = f"File {i}"
        ws["B1"] = "Test data"
        wb.save(tmp_path / f"report_{i}.xlsx")
    return tmp_path


@pytest.fixture
def nested_xlsx_dir(tmp_path):
    """Create a directory tree with xlsx files at multiple levels."""
    sub = tmp_path / "subdir"
    sub.mkdir()
    for i, d in enumerate([tmp_path, sub]):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = f"Level {i}"
        wb.save(d / f"file_{i}.xlsx")
    return tmp_path


class TestFixBatch:
    def test_fixes_all_files(self, xlsx_dir):
        results = fix_batch(xlsx_dir)
        assert len(results) == 3
        for r in results:
            assert "stats" in r
            assert r["stats"]["inline_strings"] > 0

    def test_output_dir(self, xlsx_dir, tmp_path):
        out = tmp_path / "output"
        results = fix_batch(xlsx_dir, output_dir=out)
        assert len(results) == 3
        assert out.exists()
        assert len(list(out.glob("*.xlsx"))) == 3

    def test_recursive(self, nested_xlsx_dir):
        results = fix_batch(nested_xlsx_dir, recursive=True)
        assert len(results) == 2

    def test_non_recursive_default(self, nested_xlsx_dir):
        results = fix_batch(nested_xlsx_dir, recursive=False)
        assert len(results) == 1

    def test_empty_directory(self, tmp_path):
        results = fix_batch(tmp_path)
        assert results == []

    def test_not_a_directory(self, tmp_path):
        f = tmp_path / "file.txt"
        f.write_text("hello")
        with pytest.raises(NotADirectoryError):
            fix_batch(f)

    def test_skips_temp_files(self, xlsx_dir):
        # Create a temp file (Excel lock file)
        lock = xlsx_dir / "~$report_0.xlsx"
        lock.write_bytes(b"lock")
        results = fix_batch(xlsx_dir)
        files = [r["file"].name for r in results]
        assert "~$report_0.xlsx" not in files

    def test_handles_bad_file_gracefully(self, xlsx_dir):
        bad = xlsx_dir / "bad.xlsx"
        bad.write_text("not a zip")
        results = fix_batch(xlsx_dir)
        assert len(results) == 4  # 3 good + 1 bad
        errors = [r for r in results if "error" in r]
        assert len(errors) == 1
        assert errors[0]["file"].name == "bad.xlsx"


class TestBatchCLI:
    @patch("xlsx_fixer.license.validate_license")
    def test_batch_with_valid_key(self, mock_validate, xlsx_dir, capsys):
        mock_validate.return_value = (True, "Licensed (Commercial)")
        from xlsx_fixer.cli import cmd_batch
        args = argparse.Namespace(
            directory=str(xlsx_dir),
            key="valid-key",
            key_file=None,
            output_dir=None,
            recursive=False,
        )
        exit_code = cmd_batch(args)
        assert exit_code == 0
        captured = capsys.readouterr()
        assert "3 files processed" in captured.out

    @patch("xlsx_fixer.license.validate_license")
    def test_batch_with_invalid_key(self, mock_validate, xlsx_dir, capsys):
        mock_validate.return_value = (False, "Invalid license key")
        from xlsx_fixer.cli import cmd_batch
        args = argparse.Namespace(
            directory=str(xlsx_dir),
            key="bad-key",
            key_file=None,
            output_dir=None,
            recursive=False,
        )
        exit_code = cmd_batch(args)
        assert exit_code == 1
        captured = capsys.readouterr()
        assert "License error" in captured.err

    def test_batch_no_key(self, xlsx_dir, capsys, monkeypatch):
        # Ensure no default key file exists
        monkeypatch.setattr("xlsx_fixer.license.CACHE_DIR", xlsx_dir / "no-cache")
        from xlsx_fixer.cli import cmd_batch
        args = argparse.Namespace(
            directory=str(xlsx_dir),
            key=None,
            key_file=None,
            output_dir=None,
            recursive=False,
        )
        exit_code = cmd_batch(args)
        assert exit_code == 1
        captured = capsys.readouterr()
        assert "requires a license key" in captured.err

    @patch("xlsx_fixer.license.validate_license")
    def test_batch_with_key_file(self, mock_validate, xlsx_dir, tmp_path, capsys):
        mock_validate.return_value = (True, "Licensed (Enterprise)")
        key_file = tmp_path / "my-key"
        key_file.write_text("FILE-KEY-123\n")
        from xlsx_fixer.cli import cmd_batch
        args = argparse.Namespace(
            directory=str(xlsx_dir),
            key=None,
            key_file=str(key_file),
            output_dir=None,
            recursive=False,
        )
        exit_code = cmd_batch(args)
        assert exit_code == 0
        mock_validate.assert_called_once_with("FILE-KEY-123")
